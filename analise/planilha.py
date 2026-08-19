#!/usr/bin/env python3
"""
Gera analise/acompanhamento.xlsx.

Sempre monta as abas de plano (Acompanhamento, Por serie, Como ler). Se
analise/dados/historico.json existir, acrescenta duas abas alimentadas pela API:
Coletado, com o numero mais recente de cada post real, e Curva, com a leitura dia
a dia, que e o que diz se um post seguiu rendendo ou morreu no primeiro dia.

Uso: python3 analise/planilha.py
"""
import datetime, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TINTA="FF1B1A17"; VERDE="FF33503F"; PAPEL="FFF3EFE6"; NEUTRO="FF8A8578"
PREENCHER=PatternFill("solid", fgColor="FFFFF2CC")   # amarelo: você preenche
CALC=PatternFill("solid", fgColor="FFEFEFEF")        # cinza: calculado
CAB=PatternFill("solid", fgColor=VERDE)
F="Arial"
fina=Side(style="thin", color="FFD0CCC0")
BORDA=Border(left=fina,right=fina,top=fina,bottom=fina)

wb=Workbook(); ws=wb.active; ws.title="Acompanhamento"

ws["A1"]="Acompanhamento de posts · @matheus__ramos"
ws["A1"].font=Font(F,size=14,bold=True,color=TINTA)
ws["A2"]="Preencha só as células amarelas, com os números do Insights de cada post. As cinzas se calculam sozinhas."
ws["A2"].font=Font(F,size=10,italic=True,color=NEUTRO)

ws["A4"]="Data da primeira terça:"; ws["A4"].font=Font(F,size=10,bold=True)
ws["C4"]=datetime.date(2026,8,25); ws["C4"].font=Font(F,size=10,bold=True,color="FF0000FF")
ws["C4"].fill=PREENCHER; ws["C4"].number_format="dd/mm/yyyy"; ws["C4"].border=BORDA
ws["D4"]="← troque por quando você começar. As 16 datas se ajustam."
ws["D4"].font=Font(F,size=9,italic=True,color=NEUTRO)

cabs=["#","Data","Dia","Peça","Série","Formato",
      "Alcance","Salvam.","Compart.","Coment.","Visitas perfil","Seguidores",
      "Salv./alcance","Segui./alcance","Visitas/alcance"]
L=6
for i,c in enumerate(cabs,1):
    cel=ws.cell(L,i,c); cel.font=Font(F,size=10,bold=True,color="FFFFFFFF")
    cel.fill=CAB; cel.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cel.border=BORDA
ws.row_dimensions[L].height=30

# offsets a partir da primeira terça: ter, qui, sab, dom por semana
posts=[]
series={"ter":"O Porquê do Número","qui":None,"sab":None,"dom":"O Porquê"}
plano=[
 (0,"Ter","65% de hidratação","O Porquê do Número","Carrossel"),
 (2,"Qui","Farinha importada","Opinião Impopular","Card/Reels"),
 (4,"Sáb","Criação autoral","Autoral","Foto"),
 (5,"Dom","Ponto de véu","O Porquê","Reels"),
 (7,"Ter","48h de maturação","O Porquê do Número","Carrossel"),
 (9,"Qui","24h contra 48h","Uma Variável","Carrossel"),
 (11,"Sáb","Cronograma de produção","Bastidores","Carrossel"),
 (12,"Dom","Temperatura da água","O Porquê","Reels"),
 (14,"Ter","W300, força da farinha","O Porquê do Número","Carrossel"),
 (16,"Qui","Forno caro","Opinião Impopular","Card/Reels"),
 (18,"Sáb","O erro que me custou caro","O Erro","Reels"),
 (19,"Dom","Biga contra poolish","O Porquê","Reels"),
 (21,"Ter","450°C no forno","O Porquê do Número","Carrossel"),
 (23,"Qui","60% contra 70%","Uma Variável","Carrossel"),
 (25,"Sáb","Aula, turmas em SP","Formação","Carrossel"),
 (26,"Dom","Os 3 números","O Porquê","Reels"),
]

# linha de exemplo, cinza e italico, para mostrar o formato esperado
r=L+1
ws.cell(r,1,"ex").font=Font(F,size=9,italic=True,color=NEUTRO)
for i,v in enumerate(["","","EXEMPLO (apague)","","",3200,86,24,19,140,31],2):
    c=ws.cell(r,i,v); c.font=Font(F,size=9,italic=True,color=NEUTRO); c.border=BORDA
for i,f in enumerate([f"=IFERROR(H{r}/G{r},\"\")",f"=IFERROR(L{r}/G{r},\"\")",f"=IFERROR(K{r}/G{r},\"\")"],13):
    c=ws.cell(r,i,f); c.font=Font(F,size=9,italic=True,color=NEUTRO)
    c.number_format="0.0%"; c.fill=CALC; c.border=BORDA

inicio=L+2
for n,(off,dia,peca,serie,fmt) in enumerate(plano):
    r=inicio+n
    ws.cell(r,1,n+1).font=Font(F,size=10)
    d=ws.cell(r,2,f"=$C$4+{off}"); d.number_format="dd/mm"; d.font=Font(F,size=10)
    for i,v in enumerate([dia,peca,serie,fmt],3):
        ws.cell(r,i,v).font=Font(F,size=10)
    for i in range(7,13):                      # G..L amarelo, você preenche
        c=ws.cell(r,i); c.fill=PREENCHER; c.font=Font(F,size=10); c.border=BORDA
        c.number_format="#,##0"
    for i,f in enumerate([f"=IFERROR(H{r}/G{r},\"\")",f"=IFERROR(L{r}/G{r},\"\")",f"=IFERROR(K{r}/G{r},\"\")"],13):
        c=ws.cell(r,i,f); c.number_format="0.0%"; c.fill=CALC
        c.font=Font(F,size=10); c.border=BORDA
    for i in range(1,7): ws.cell(r,i).border=BORDA

fim=inicio+len(plano)-1
rt=fim+2
ws.cell(rt,4,"MÉDIA DO MÊS").font=Font(F,size=10,bold=True)
for i,col in zip(range(7,13),"GHIJKL"):
    c=ws.cell(rt,i,f"=IFERROR(AVERAGE({col}{inicio}:{col}{fim}),\"\")")
    c.font=Font(F,size=10,bold=True); c.number_format="#,##0"; c.border=BORDA
for i,col in zip(range(13,16),"MNO"):
    c=ws.cell(rt,i,f"=IFERROR(AVERAGE({col}{inicio}:{col}{fim}),\"\")")
    c.font=Font(F,size=10,bold=True); c.number_format="0.0%"; c.border=BORDA

rb=rt+1
ws.cell(rb,4,"Ponto de partida (agosto/26)").font=Font(F,size=10,italic=True,color=NEUTRO)
ws.cell(rb,12,14).font=Font(F,size=10,italic=True,color=NEUTRO)     # 14 seguidores/post
ws.cell(rb,14,0.003).font=Font(F,size=10,italic=True,color=NEUTRO)
ws.cell(rb,14).number_format="0.0%"
ws.cell(rb,15,"← a meta é dobrar esta coluna").font=Font(F,size=9,italic=True,color=NEUTRO)

larg=[5,10,7,30,22,13,11,10,10,10,14,12,14,15,15]
for i,w in enumerate(larg,1): ws.column_dimensions[get_column_letter(i)].width=w
ws.freeze_panes="D7"

# ---------- por série ----------
s2=wb.create_sheet("Por série")
s2["A1"]="Qual série retém mais"; s2["A1"].font=Font(F,size=14,bold=True,color=TINTA)
s2["A2"]="Tudo aqui se calcula da aba Acompanhamento. Não preencha nada."
s2["A2"].font=Font(F,size=10,italic=True,color=NEUTRO)
cab2=["Série","Posts","Alcance médio","Salv./alcance","Segui./alcance","Seguidores no total"]
for i,c in enumerate(cab2,1):
    cel=s2.cell(4,i,c); cel.font=Font(F,size=10,bold=True,color="FFFFFFFF")
    cel.fill=CAB; cel.alignment=Alignment(horizontal="center",wrap_text=True); cel.border=BORDA
s2.row_dimensions[4].height=28
A=f"Acompanhamento!$E${inicio}:$E${fim}"
for n,serie in enumerate(["O Porquê do Número","Opinião Impopular","Uma Variável","O Porquê","Autoral","Bastidores","O Erro","Formação"]):
    r=5+n
    s2.cell(r,1,serie).font=Font(F,size=10)
    s2.cell(r,2,f'=COUNTIF({A},A{r})').font=Font(F,size=10)
    s2.cell(r,3,f'=IFERROR(AVERAGEIFS(Acompanhamento!$G${inicio}:$G${fim},{A},A{r}),"")').number_format="#,##0"
    s2.cell(r,4,f'=IFERROR(SUMIFS(Acompanhamento!$H${inicio}:$H${fim},{A},A{r})/SUMIFS(Acompanhamento!$G${inicio}:$G${fim},{A},A{r}),"")').number_format="0.0%"
    s2.cell(r,5,f'=IFERROR(SUMIFS(Acompanhamento!$L${inicio}:$L${fim},{A},A{r})/SUMIFS(Acompanhamento!$G${inicio}:$G${fim},{A},A{r}),"")').number_format="0.0%"
    s2.cell(r,6,f'=IFERROR(SUMIFS(Acompanhamento!$L${inicio}:$L${fim},{A},A{r}),"")').number_format="#,##0"
    for i in range(2,7):
        s2.cell(r,i).font=Font(F,size=10); s2.cell(r,i).fill=CALC; s2.cell(r,i).border=BORDA
    s2.cell(r,1).border=BORDA
for i,w in enumerate([24,8,15,15,16,20],1): s2.column_dimensions[get_column_letter(i)].width=w

# ---------- como ler ----------
s3=wb.create_sheet("Como ler")
s3["A1"]="Como ler os números"; s3["A1"].font=Font(F,size=14,bold=True,color=TINTA)
linhas=[
 ("",""),
 ("Salv./alcance","O indicador mais importante aqui. Conteúdo educativo bom é salvo, e salvamento é o sinal mais forte que existe pro algoritmo reentregar."),
 ("","Abaixo de 1%: o post informou mas não foi útil o suficiente pra guardar."),
 ("","1% a 3%: bom. É a faixa que sustenta crescimento."),
 ("","Acima de 3%: repita o formato. Achou uma veia."),
 ("",""),
 ("Segui./alcance","A conversão. É o número que responde se o post fez a pessoa querer ficar."),
 ("","Hoje o perfil está abaixo de 0,3%, calculado sobre 49.900 views e o histórico de 14 seguidores por post."),
 ("","A meta dos 30 dias é dobrar isso. Chegando em 1% com o mesmo alcance de hoje, são cerca de 500 seguidores por mês."),
 ("",""),
 ("Visitas/alcance","Intenção. A pessoa viu o post e foi conferir o perfil."),
 ("","Visita alta com conversão baixa significa que o post funciona e o perfil não segura. Aí o problema é bio e fixados, não conteúdo."),
 ("",""),
 ("Comentários","Não entra no cálculo de retenção, mas é o que mais empurra alcance. Espere Opinião Impopular liderar aqui."),
 ("",""),
 ("Onde achar","Instagram, abre o post, toca em Ver insights. Reels mostra também tempo médio de visualização, que vale anotar à parte."),
 ("",""),
 ("Quando analisar","Espere 72 horas depois de publicar antes de anotar, porque o post continua rodando. Compare só depois de 8 posts, que é quando o padrão aparece."),
 ("",""),
 ("O que a análise decide","No fim das 4 semanas, a aba Por série mostra qual formato retém mais. O mês seguinte se escreve em cima dessa resposta, não de palpite."),
]
r=3
for a,b in linhas:
    if a: s3.cell(r,1,a).font=Font(F,size=10,bold=True,color=VERDE)
    if b: s3.cell(r,2,b).font=Font(F,size=10)
    r+=1
s3.column_dimensions["A"].width=20; s3.column_dimensions["B"].width=115
for rr in range(3,r):
    s3.cell(rr,2).alignment=Alignment(wrap_text=True,vertical="top")


# ---------- abas alimentadas pela API ----------
BASE = Path(__file__).resolve().parent
HIST = BASE / "dados" / "historico.json"

def taxa(c):
    c.number_format="0.0%"; c.font=Font(F,size=10); c.fill=CALC; c.border=BORDA

if HIST.exists():
    hist=json.loads(HIST.read_text()).get("leituras",[])
else:
    hist=[]

if hist:
    ultima=hist[-1]
    posts=sorted(ultima["posts"].items(),
                 key=lambda kv: kv[1].get("publicado_em") or "", reverse=True)

    s4=wb.create_sheet("Coletado", 1)
    s4["A1"]="Coletado da API"; s4["A1"].font=Font(F,size=14,bold=True,color=TINTA)
    s4["A2"]=f'Ultima leitura: {ultima["lido_em"]}. Gerado por analise/planilha.py, nao edite a mao.'
    s4["A2"].font=Font(F,size=10,italic=True,color=NEUTRO)
    ind=ultima.get("metricas_indisponiveis") or []
    if ind:
        s4["A3"]=f'Metricas que a API nao entregou: {", ".join(ind)}'
        s4["A3"].font=Font(F,size=10,italic=True,color="FFB00000")

    cab4=["Publicado","Tipo","Legenda","Alcance","Salvam.","Compart.","Coment.",
          "Visitas perfil","Seguidores","Salv./alcance","Segui./alcance","Link"]
    for i,c in enumerate(cab4,1):
        cel=s4.cell(5,i,c); cel.font=Font(F,size=10,bold=True,color="FFFFFFFF")
        cel.fill=CAB; cel.alignment=Alignment(horizontal="center",wrap_text=True); cel.border=BORDA
    s4.row_dimensions[5].height=28

    for n,(mid,p) in enumerate(posts):
        r=6+n; m=p.get("metricas",{})
        pub=(p.get("publicado_em") or "")[:10]
        s4.cell(r,1,pub).font=Font(F,size=10)
        s4.cell(r,2,p.get("tipo","")).font=Font(F,size=10)
        s4.cell(r,3,p.get("legenda","")).font=Font(F,size=10)
        for i,k in enumerate(["reach","saved","shares","comments","profile_visits","follows"],4):
            c=s4.cell(r,i,m.get(k)); c.font=Font(F,size=10)
            c.number_format="#,##0"; c.fill=CALC; c.border=BORDA
        taxa(s4.cell(r,10,f'=IFERROR(E{r}/D{r},"")'))
        taxa(s4.cell(r,11,f'=IFERROR(I{r}/D{r},"")'))
        c=s4.cell(r,12,p.get("permalink","")); c.font=Font(F,size=9,color="FF0000FF")
        for i in (1,2,3): s4.cell(r,i).border=BORDA
    for i,w in enumerate([12,10,42,10,10,10,10,13,12,14,15,34],1):
        s4.column_dimensions[get_column_letter(i)].width=w
    s4.freeze_panes="D6"

    # ---------- curva ----------
    s5=wb.create_sheet("Curva", 2)
    s5["A1"]="Curva de salvamento, dia a dia"
    s5["A1"].font=Font(F,size=14,bold=True,color=TINTA)
    s5["A2"]="Cada linha e um post, cada coluna uma leitura. Numero que para de subir e post que morreu."
    s5["A2"].font=Font(F,size=10,italic=True,color=NEUTRO)
    dias=[l["lido_em"][:10] for l in hist]
    s5.cell(4,1,"Post").font=Font(F,size=10,bold=True,color="FFFFFFFF")
    s5.cell(4,1).fill=CAB; s5.cell(4,1).border=BORDA
    for j,d in enumerate(dias):
        c=s5.cell(4,2+j,d); c.font=Font(F,size=9,bold=True,color="FFFFFFFF")
        c.fill=CAB; c.alignment=Alignment(horizontal="center"); c.border=BORDA
    for n,(mid,p) in enumerate(posts):
        r=5+n
        s5.cell(r,1,p.get("legenda","")[:44]).font=Font(F,size=10)
        s5.cell(r,1).border=BORDA
        for j,l in enumerate(hist):
            v=l["posts"].get(mid,{}).get("metricas",{}).get("saved")
            c=s5.cell(r,2+j,v); c.font=Font(F,size=10)
            c.number_format="#,##0"; c.fill=CALC; c.border=BORDA
    s5.column_dimensions["A"].width=46
    for j in range(len(dias)): s5.column_dimensions[get_column_letter(2+j)].width=11
    s5.freeze_panes="B5"

saida = BASE / "acompanhamento.xlsx"
wb.save(saida)
print(f"salvo: {saida}" + (f" · {len(hist)} leituras no historico" if hist else " · sem historico ainda"))
